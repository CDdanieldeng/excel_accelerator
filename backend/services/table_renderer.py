"""Table rendering service for converting sheet data to PNG images."""

import csv
import logging
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

logger = logging.getLogger(__name__)

# Type aliases
CellValue = Optional[str]
Row = List[CellValue]
Grid = List[Row]


class UnsupportedFileTypeError(Exception):
    """Raised when file type is not supported."""

    pass


def get_sheet_list(file_path: str) -> Tuple[str, List[str]]:
    """
    Get list of sheet names from a file.

    Args:
        file_path: Path to the file

    Returns:
        Tuple of (file_type, list of sheet names)

    Raises:
        UnsupportedFileTypeError: If file type is not supported
    """
    file_ext = Path(file_path).suffix.lower().lstrip(".")
    logger.info(f"Getting sheet list: file={file_path}")

    if file_ext == "xlsx":
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            logger.info(f"Found {len(sheet_names)} sheets in XLSX file")
            return (file_ext, sheet_names)
        except Exception as e:
            logger.exception(f"Error reading XLSX file: {file_path}")
            raise

    elif file_ext == "csv":
        # CSV files have a single default sheet
        logger.info("CSV file has single default sheet")
        return (file_ext, ["__default__"])

    elif file_ext == "xlsb":
        raise NotImplementedError("xlsb format is not yet supported for sheet list")
    else:
        raise UnsupportedFileTypeError(f"Unsupported file type: {file_ext}")


def load_sheet_window(
    file_path: str,
    sheet_name: str,
    row_start: int,
    row_end: int,
    col_start: int,
    col_end: int,
    max_rows: int = 200,
    max_cols: int = 50,
) -> Grid:
    """
    Load a data window from the specified sheet.

    Args:
        file_path: Path to the file
        sheet_name: Name of the sheet (use "__default__" for CSV)
        row_start: 0-based start row index (inclusive)
        row_end: 0-based end row index (inclusive)
        col_start: 0-based start column index (inclusive)
        col_end: 0-based end column index (inclusive)
        max_rows: Maximum number of rows to load
        max_cols: Maximum number of columns to load

    Returns:
        Grid: 2D array where grid[row][col] is the cell value

    Raises:
        UnsupportedFileTypeError: If file type is not supported
    """
    file_ext = Path(file_path).suffix.lower().lstrip(".")
    logger.info(
        f"Loading sheet window: file={file_path}, sheet={sheet_name}, "
        f"rows=[{row_start}, {row_end}], cols=[{col_start}, {col_end}]"
    )

    # Validate and clamp ranges
    actual_row_start = max(0, row_start)
    actual_row_end = min(row_end, row_start + max_rows - 1)
    actual_col_start = max(0, col_start)
    actual_col_end = min(col_end, col_start + max_cols - 1)

    if file_ext == "xlsx":
        return _load_xlsx_window(
            file_path, sheet_name, actual_row_start, actual_row_end, actual_col_start, actual_col_end
        )
    elif file_ext == "csv":
        return _load_csv_window(
            file_path, actual_row_start, actual_row_end, actual_col_start, actual_col_end
        )
    elif file_ext == "xlsb":
        raise NotImplementedError("xlsb format is not yet supported for sheet image rendering")
    else:
        raise UnsupportedFileTypeError(f"Unsupported file type: {file_ext}")


def _load_xlsx_window(
    file_path: str,
    sheet_name: str,
    row_start: int,
    row_end: int,
    col_start: int,
    col_end: int,
) -> Grid:
    """Load window from XLSX file."""
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)

        if sheet_name not in workbook.sheetnames:
            workbook.close()
            raise ValueError(f"Sheet '{sheet_name}' not found in file")

        sheet = workbook[sheet_name]
        grid: Grid = []

        # Iterate through rows
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_idx < row_start:
                continue
            if row_idx > row_end:
                break

            # Extract cells in the column range
            row_data: Row = []
            for col_idx in range(col_start, col_end + 1):
                if col_idx < len(row):
                    cell_value = row[col_idx]
                    # Convert to string or None
                    if cell_value is None:
                        row_data.append(None)
                    else:
                        row_data.append(str(cell_value))
                else:
                    row_data.append(None)

            grid.append(row_data)

        workbook.close()

        logger.info(
            f"Loaded XLSX window: {len(grid)} rows, "
            f"{max(len(row) for row in grid) if grid else 0} columns"
        )

        return grid

    except Exception as e:
        logger.exception(f"Error loading XLSX window: {file_path}, sheet={sheet_name}")
        raise


def _load_csv_window(
    file_path: str,
    row_start: int,
    row_end: int,
    col_start: int,
    col_end: int,
) -> Grid:
    """Load window from CSV file."""
    try:
        grid: Grid = []

        with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader):
                if row_idx < row_start:
                    continue
                if row_idx > row_end:
                    break

                # Extract cells in the column range
                row_data: Row = []
                for col_idx in range(col_start, col_end + 1):
                    if col_idx < len(row):
                        cell_value = row[col_idx]
                        # Convert to string or None
                        if cell_value and cell_value.strip():
                            row_data.append(cell_value.strip())
                        else:
                            row_data.append(None)
                    else:
                        row_data.append(None)

                grid.append(row_data)

        logger.info(
            f"Loaded CSV window: {len(grid)} rows, "
            f"{max(len(row) for row in grid) if grid else 0} columns"
        )

        return grid

    except Exception as e:
        logger.exception(f"Error loading CSV window: {file_path}")
        raise


class TableImageRenderer:
    """Renders a grid of data as a PNG image."""

    def __init__(
        self,
        row_height_px: int = 44,  # Increased for higher resolution (was 22)
        col_width_px: int = 180,  # Increased for higher resolution (was 90)
        font_size: int = 24,  # Increased for higher resolution (was 12)
        padding_px: int = 8,  # Increased for higher resolution (was 4)
        scale_factor: float = 2.0,  # Scale factor for high DPI rendering
    ) -> None:
        """
        Initialize the table image renderer.

        Args:
            row_height_px: Height of each row in pixels (base size, will be scaled)
            col_width_px: Width of each column in pixels (base size, will be scaled)
            font_size: Font size for text (base size, will be scaled)
            padding_px: Padding inside each cell (base size, will be scaled)
            scale_factor: Scale factor for high DPI rendering (default 2.0 for 2x resolution)
        """
        self.scale_factor = scale_factor
        # Apply scale factor to all dimensions
        self.row_height_px = int(row_height_px * scale_factor)
        self.col_width_px = int(col_width_px * scale_factor)
        self.font_size = int(font_size * scale_factor)
        self.padding_px = int(padding_px * scale_factor)

        # Try to load a font that supports Chinese characters (UTF-8)
        # Priority: Chinese fonts > system fonts > default
        self.font = None
        
        # Try Chinese fonts first (macOS)
        # Note: .ttc files may contain multiple fonts, we use index 0 by default
        chinese_font_paths = [
            ("/System/Library/Fonts/PingFang.ttc", 0),  # macOS Chinese font
            ("/System/Library/Fonts/STHeiti Light.ttc", 0),  # macOS Chinese font (alternative)
            ("/System/Library/Fonts/STSong.ttc", 0),  # macOS Chinese font (alternative)
            ("/System/Library/Fonts/Supplemental/Songti.ttc", 0),  # macOS Chinese font (alternative)
        ]
        
        for font_path, font_index in chinese_font_paths:
            try:
                self.font = ImageFont.truetype(font_path, self.font_size, index=font_index)
                logger.info(f"Loaded Chinese font: {font_path} at size {self.font_size}")
                break
            except (OSError, IOError, TypeError):
                # TypeError may occur if index parameter is not supported
                try:
                    self.font = ImageFont.truetype(font_path, self.font_size)
                    logger.info(f"Loaded Chinese font: {font_path} at size {self.font_size}")
                    break
                except (OSError, IOError):
                    continue
        
        # If Chinese fonts not found, try system fonts
        if self.font is None:
            system_font_paths = [
                "/System/Library/Fonts/Helvetica.ttc",  # macOS
                "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",  # Linux (Noto Sans CJK)
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",  # Linux (fallback)
                "C:/Windows/Fonts/msyh.ttc",  # Windows (Microsoft YaHei)
                "C:/Windows/Fonts/simsun.ttc",  # Windows (SimSun)
            ]
            
            for font_path in system_font_paths:
                try:
                    self.font = ImageFont.truetype(font_path, self.font_size)
                    logger.info(f"Loaded system font: {font_path} at size {self.font_size}")
                    break
                except (OSError, IOError):
                    continue
        
        # Fall back to default font (may not support Chinese)
        if self.font is None:
            logger.warning("Could not load custom font, using default (may not support Chinese)")
            self.font = ImageFont.load_default()

    def _truncate_text_to_fit(
        self,
        text: str,
        available_width: int,
        draw: ImageDraw.Draw,
    ) -> str:
        """
        Truncate text to fit within available width, adding ellipsis if needed.

        Args:
            text: Text to truncate
            available_width: Available width in pixels
            draw: ImageDraw object for measuring text

        Returns:
            Truncated text with ellipsis if needed
        """
        if not text:
            return ""

        # Measure ellipsis width
        ellipsis = "…"
        try:
            ellipsis_bbox = draw.textbbox((0, 0), ellipsis, font=self.font)
            ellipsis_width = ellipsis_bbox[2] - ellipsis_bbox[0]
        except Exception:
            ellipsis = "..."
            ellipsis_bbox = draw.textbbox((0, 0), ellipsis, font=self.font)
            ellipsis_width = ellipsis_bbox[2] - ellipsis_bbox[0]

        # If ellipsis itself doesn't fit, return it anyway
        if ellipsis_width >= available_width:
            return ellipsis

        # Measure full text width
        try:
            full_bbox = draw.textbbox((0, 0), text, font=self.font)
            full_width = full_bbox[2] - full_bbox[0]
        except Exception:
            # If measurement fails, use character count as fallback
            max_chars = max(1, available_width // (self.font_size // 2))
            if len(text) > max_chars:
                return text[:max_chars] + ellipsis
            return text

        # If text fits, return as is
        if full_width <= available_width:
            return text

        # Binary search for the right truncation point
        # Start with approximate truncation based on width ratio
        approx_chars = max(1, int(len(text) * (available_width - ellipsis_width) / full_width))
        low = 0
        high = len(text)
        best_length = 0

        # Binary search to find maximum length that fits
        while low <= high:
            mid = (low + high) // 2
            if mid == 0:
                break
            
            truncated = text[:mid]
            try:
                bbox = draw.textbbox((0, 0), truncated, font=self.font)
                width = bbox[2] - bbox[0]
                total_width = width + ellipsis_width
                
                if total_width <= available_width:
                    # This length fits, try longer
                    best_length = mid
                    low = mid + 1
                else:
                    # Too long, try shorter
                    high = mid - 1
            except Exception:
                # If measurement fails, use character count approximation
                high = mid - 1
                continue

        # If no length fits, return ellipsis only
        if best_length == 0:
            return ellipsis

        return text[:best_length] + ellipsis

    def render_grid(
        self,
        grid: Grid,
        row_offset: int,
        col_offset: int,
    ) -> Tuple[bytes, int, int]:
        """
        Render a grid as a PNG image.

        Args:
            grid: 2D array where grid[row][col] is the cell text
            row_offset: Starting row index in the real sheet (0-based)
            col_offset: Starting column index in the real sheet (0-based)

        Returns:
            Tuple of (png_bytes, row_height_px, col_width_px)
        """
        if not grid:
            # Empty grid - create a minimal image
            grid = [[None]]

        rows = len(grid)
        cols = max(len(row) for row in grid) if grid else 1

        # Add one row for column headers, one column for row numbers
        total_rows = rows + 1
        total_cols = cols + 1

        # Calculate image dimensions
        width = total_cols * self.col_width_px
        height = total_rows * self.row_height_px

        logger.info(
            f"Rendering grid: {rows} rows x {cols} cols, "
            f"image size: {width}x{height} pixels (scale: {self.scale_factor}x)"
        )

        # Create high-resolution image
        image = Image.new("RGB", (width, height), color="white")
        # Use standard RGB mode for drawing (RGBA mode has issues with some operations)
        draw = ImageDraw.Draw(image)

        # Draw grid lines and content
        for row_idx in range(total_rows):
            for col_idx in range(total_cols):
                x1 = col_idx * self.col_width_px
                y1 = row_idx * self.row_height_px
                x2 = x1 + self.col_width_px
                y2 = y1 + self.row_height_px

                # Draw cell border (thicker for high resolution)
                border_width = max(1, int(self.scale_factor))
                draw.rectangle([x1, y1, x2 - 1, y2 - 1], outline="gray", width=border_width)

                # Determine cell content
                if row_idx == 0:
                    # Column header row
                    if col_idx == 0:
                        # Top-left corner (empty)
                        text = ""
                    else:
                        # Column number (1-based for user)
                        text = str(col_offset + col_idx)
                elif col_idx == 0:
                    # Row number column
                    # Row number (1-based for user)
                    text = str(row_offset + row_idx)
                else:
                    # Data cell
                    data_row_idx = row_idx - 1
                    data_col_idx = col_idx - 1

                    if data_row_idx < len(grid) and data_col_idx < len(grid[data_row_idx]):
                        cell_value = grid[data_row_idx][data_col_idx]
                        # Ensure text is properly encoded as UTF-8 string
                        if cell_value is not None:
                            if isinstance(cell_value, bytes):
                                text = cell_value.decode('utf-8', errors='replace')
                            else:
                                text = str(cell_value)
                        else:
                            text = ""
                    else:
                        text = ""

                # Truncate text to fit cell width before rendering
                if text:
                    # Calculate available width (cell width minus padding on both sides)
                    available_width = self.col_width_px - (self.padding_px * 2)
                    
                    # Truncate text to fit available width
                    truncated_text = self._truncate_text_to_fit(text, available_width, draw)
                    
                    try:
                        # Get text bounding box for truncated text
                        bbox = draw.textbbox((0, 0), truncated_text, font=self.font)
                        text_width = bbox[2] - bbox[0]
                        text_height = bbox[3] - bbox[1]

                        # Calculate text position (centered)
                        text_x = x1 + (self.col_width_px - text_width) / 2
                        text_y = y1 + (self.row_height_px - text_height) / 2

                        # Draw truncated text
                        draw.text(
                            (text_x, text_y),
                            truncated_text,
                            fill="black",
                            font=self.font,
                        )
                    except Exception as e:
                        # If text rendering fails, log and skip
                        logger.warning(f"Failed to render text '{truncated_text[:20]}...': {e}")

        # Convert to PNG bytes with high quality settings
        png_buffer = BytesIO()
        # Save with optimization level 0 (no compression) for maximum quality
        # Higher compression levels can reduce quality slightly
        image.save(
            png_buffer,
            format="PNG",
            optimize=False,  # Disable optimization for faster saving
            compress_level=1,  # Low compression for better quality
        )
        png_bytes = png_buffer.getvalue()
        png_buffer.close()

        logger.info(f"Rendered PNG: {len(png_bytes)} bytes at {width}x{height} resolution")

        # Return the actual pixel dimensions used (scaled)
        return (png_bytes, self.row_height_px, self.col_width_px)

