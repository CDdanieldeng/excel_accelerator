"""File loading and sampling service for various file formats."""

import csv
import logging
from pathlib import Path
from typing import List, Optional, Protocol

from openpyxl import load_workbook

from backend.config import DEFAULT_MAX_SCAN_ROWS

logger = logging.getLogger(__name__)

try:
    import pyxlsb
except ImportError:
    pyxlsb = None
    logger.warning("pyxlsb not available, xlsb files will not be supported")


class SheetSample(Protocol):
    """Protocol for sheet sample data."""

    name: str
    rows: List[List[Optional[str]]]


class SheetSampleImpl:
    """Implementation of SheetSample."""

    def __init__(self, name: str, rows: List[List[Optional[str]]]):
        self.name = name
        self.rows = rows


def load_csv_sample(file_path: str, max_scan_rows: int) -> List[SheetSample]:
    """
    Load and sample a CSV file.

    Args:
        file_path: Path to the CSV file
        max_scan_rows: Maximum number of rows to scan

    Returns:
        List of SheetSample (CSV is treated as a single sheet)
    """
    logger.info(f"Loading CSV file: {file_path}, max_rows={max_scan_rows}")
    rows: List[List[Optional[str]]] = []

    try:
        with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i >= max_scan_rows:
                    break
                # Convert to List[Optional[str]]
                normalized_row: List[Optional[str]] = [
                    cell.strip() if cell and cell.strip() else None for cell in row
                ]
                rows.append(normalized_row)

        logger.info(f"Loaded {len(rows)} rows from CSV")
        return [SheetSampleImpl(name="__default__", rows=rows)]

    except Exception as e:
        logger.exception(f"Error loading CSV file: {file_path}")
        raise


def load_xlsx_sample(file_path: str, max_scan_rows: int) -> List[SheetSample]:
    """
    Load and sample an XLSX file.

    Args:
        file_path: Path to the XLSX file
        max_scan_rows: Maximum number of rows to scan per sheet

    Returns:
        List of SheetSample, one per sheet
    """
    logger.info(f"Loading XLSX file: {file_path}, max_rows={max_scan_rows}")
    samples: List[SheetSample] = []

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            logger.info(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            rows: List[List[Optional[str]]] = []

            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= max_scan_rows:
                    break
                normalized_row: List[Optional[str]] = [
                    str(cell).strip() if cell is not None and str(cell).strip() else None
                    for cell in row
                ]
                rows.append(normalized_row)

            logger.info(f"Loaded {len(rows)} rows from sheet: {sheet_name}")
            samples.append(SheetSampleImpl(name=sheet_name, rows=rows))

        workbook.close()
        return samples

    except Exception as e:
        logger.exception(f"Error loading XLSX file: {file_path}")
        raise


def load_xlsb_sample(file_path: str, max_scan_rows: int) -> List[SheetSample]:
    """
    Load and sample an XLSB file.

    Args:
        file_path: Path to the XLSB file
        max_scan_rows: Maximum number of rows to scan per sheet

    Returns:
        List of SheetSample, one per sheet
    """
    if pyxlsb is None:
        raise ValueError("pyxlsb is not installed, cannot read xlsb files")

    logger.info(f"Loading XLSB file: {file_path}, max_rows={max_scan_rows}")
    samples: List[SheetSample] = []

    try:
        with pyxlsb.open_workbook(file_path) as workbook:
            for sheet_name in workbook.sheets:
                logger.info(f"Processing sheet: {sheet_name}")
                rows: List[List[Optional[str]]] = []

                with workbook.get_sheet(sheet_name) as sheet:
                    for i, row in enumerate(sheet.rows()):
                        if i >= max_scan_rows:
                            break
                        normalized_row: List[Optional[str]] = [
                            str(cell.v).strip() if cell.v is not None and str(cell.v).strip() else None
                            for cell in row
                        ]
                        rows.append(normalized_row)

                logger.info(f"Loaded {len(rows)} rows from sheet: {sheet_name}")
                samples.append(SheetSampleImpl(name=sheet_name, rows=rows))

        return samples

    except Exception as e:
        logger.exception(f"Error loading XLSB file: {file_path}")
        raise


def load_file_sample(
    file_path: str,
    file_type: str,
    max_scan_rows: int = DEFAULT_MAX_SCAN_ROWS,
) -> List[SheetSample]:
    """
    Load and sample a file based on its type.

    Args:
        file_path: Path to the file
        file_type: File type (xlsx, csv, xlsb)
        max_scan_rows: Maximum number of rows to scan

    Returns:
        List of SheetSample objects
    """
    file_type_lower = file_type.lower()

    if file_type_lower == "csv":
        return load_csv_sample(file_path, max_scan_rows)
    elif file_type_lower == "xlsx":
        return load_xlsx_sample(file_path, max_scan_rows)
    elif file_type_lower == "xlsb":
        return load_xlsb_sample(file_path, max_scan_rows)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")

