"""Service for building pandas DataFrame from sheet data with specified header row."""

import logging
import uuid
from pathlib import Path
from typing import Any, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _clean_cell_value(cell_value: Optional[str]) -> Optional[Any]:
    """
    Clean and convert a single cell value.

    Args:
        cell_value: Raw cell value from sheet

    Returns:
        Cleaned and converted value (int, float, str, or None)
    """
    if cell_value is None:
        return None

    # Convert to string and strip whitespace
    cell_str = str(cell_value).strip()

    # Handle empty strings
    if not cell_str:
        return None

    # Excel convention: "-" represents 0
    if cell_str == "-":
        return 0

    # Try to convert to number
    try:
        # Try int first (if no decimal point)
        if "." not in cell_str and "e" not in cell_str.lower() and "E" not in cell_str:
            return int(cell_str)
        else:
            return float(cell_str)
    except ValueError:
        # Not a number, return cleaned string
        return cell_str


def _preprocess_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Preprocess DataFrame: remove empty columns and ensure proper data types.

    Args:
        df: Input DataFrame

    Returns:
        Preprocessed DataFrame
    """
    # Remove columns that are completely empty (all NaN/None)
    original_cols = len(df.columns)
    df = df.dropna(axis=1, how="all")
    removed_cols = original_cols - len(df.columns)
    if removed_cols > 0:
        logger.info(f"Removed {removed_cols} completely empty columns")

    # Convert numeric columns to proper types (avoid object dtype for numbers)
    for col in df.columns:
        # Check if column can be converted to numeric
        if df[col].dtype == "object":
            # Try to convert to numeric
            numeric_series = pd.to_numeric(df[col], errors="coerce")
            # If most values are numeric, convert the column
            if numeric_series.notna().sum() > len(df) * 0.5:  # More than 50% are numeric
                df[col] = numeric_series
                logger.debug(f"Converted column '{col}' to numeric type")

    return df


def load_full_sheet(file_path: str, sheet_name: str) -> List[List[Optional[str]]]:
    """
    Load full sheet data as a 2D array.

    Args:
        file_path: Path to the file
        sheet_name: Name of the sheet (use "__default__" for CSV)

    Returns:
        2D array where grid[row][col] is the cell value

    Raises:
        ValueError: If sheet not found or file read error
    """
    file_ext = Path(file_path).suffix.lower().lstrip(".")
    logger.info(f"Loading full sheet: file={file_path}, sheet={sheet_name}")

    if file_ext == "xlsx":
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            if sheet_name not in workbook.sheetnames:
                workbook.close()
                raise ValueError(f"Sheet '{sheet_name}' not found in file")

            sheet = workbook[sheet_name]
            grid: List[List[Optional[str]]] = []

            # Read all rows
            for row in sheet.iter_rows(values_only=True):
                # Convert to list of strings or None
                row_data: List[Optional[str]] = []
                for cell_value in row:
                    if cell_value is None:
                        row_data.append(None)
                    else:
                        row_data.append(str(cell_value))
                grid.append(row_data)

            workbook.close()

            logger.info(f"Loaded {len(grid)} rows from XLSX sheet")
            return grid

        except Exception as e:
            logger.exception(f"Error loading XLSX sheet: {file_path}, sheet={sheet_name}")
            raise

    elif file_ext == "csv":
        try:
            grid: List[List[Optional[str]]] = []
            import csv

            with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
                reader = csv.reader(f)
                for row in reader:
                    # Convert to list of strings or None
                    row_data: List[Optional[str]] = []
                    for cell_value in row:
                        if cell_value and cell_value.strip():
                            row_data.append(cell_value.strip())
                        else:
                            row_data.append(None)
                    grid.append(row_data)

            logger.info(f"Loaded {len(grid)} rows from CSV file")
            return grid

        except Exception as e:
            logger.exception(f"Error loading CSV file: {file_path}")
            raise

    else:
        raise ValueError(f"Unsupported file type: {file_ext}")


def build_dataframe_from_header(
    file_path: str,
    sheet_name: str,
    header_row_number: int,  # 1-based
    max_preview_rows: int = 100,
) -> Tuple[str, pd.DataFrame, List[List[Optional[Any]]]]:
    """
    Build pandas DataFrame from sheet data using specified header row.

    Args:
        file_path: Path to the file
        sheet_name: Name of the sheet
        header_row_number: Header row number (1-based)
        max_preview_rows: Maximum number of preview rows to return

    Returns:
        Tuple of (dataset_id, DataFrame, preview_rows)

    Raises:
        ValueError: If header row is invalid or empty
    """
    logger.info(
        f"Building DataFrame: file={file_path}, sheet={sheet_name}, "
        f"header_row={header_row_number} (1-based)"
    )

    # Convert 1-based to 0-based
    header_row_index = header_row_number - 1

    if header_row_index < 0:
        raise ValueError(f"Header row number must be >= 1, got {header_row_number}")

    # Load full sheet
    grid = load_full_sheet(file_path, sheet_name)

    if len(grid) == 0:
        raise ValueError("Sheet is empty")

    # Validate header row index
    if header_row_index >= len(grid):
        raise ValueError(
            f"Header row {header_row_number} is out of range. "
            f"Sheet has {len(grid)} rows (1-based: row 1 to {len(grid)})"
        )

    # Extract header row
    header_row = grid[header_row_index]
    if not header_row:
        raise ValueError(f"Header row {header_row_number} is empty")

    # Auto-detect actual column count: find the rightmost non-empty cell
    # This handles cases where header row has trailing empty cells
    actual_col_count = 0
    for i in range(len(header_row) - 1, -1, -1):
        if header_row[i] and str(header_row[i]).strip():
            actual_col_count = i + 1
            break

    if actual_col_count == 0:
        raise ValueError(f"Header row {header_row_number} is completely empty")

    # Process column names (only up to actual_col_count)
    columns: List[str] = []
    col_index = 1

    for i in range(actual_col_count):
        cell_value = header_row[i] if i < len(header_row) else None
        if cell_value and str(cell_value).strip():
            col_name = str(cell_value).strip()
            columns.append(col_name)
        else:
            # Empty column name, auto-fill
            columns.append(f"col_{col_index}")
            col_index += 1

    logger.info(f"Extracted {len(columns)} columns from header row (detected {actual_col_count} actual columns)")

    # Extract data rows (from header_row_index + 1 to end)
    data_start_index = header_row_index + 1
    data_rows: List[List[Optional[Any]]] = []

    # Use actual column count
    max_cols = len(columns)
    for row in grid[data_start_index:]:
        # Check if row is completely empty
        if not any(cell and str(cell).strip() for cell in row):
            continue  # Skip completely empty rows

        # Process row: pad with None if too short, truncate if too long
        row_data: List[Optional[Any]] = []
        for i in range(max_cols):
            if i < len(row):
                cell_value = row[i]
                # Clean and convert cell value
                cleaned_value = _clean_cell_value(cell_value)
                row_data.append(cleaned_value)
            else:
                row_data.append(None)

        data_rows.append(row_data)

    logger.info(f"Extracted {len(data_rows)} data rows (skipped empty rows)")

    # Build DataFrame
    try:
        df = pd.DataFrame(data_rows, columns=columns)
        logger.info(f"Built DataFrame (before preprocessing): shape={df.shape}")

        # Preprocess DataFrame: remove completely empty columns
        df = _preprocess_dataframe(df)
        logger.info(f"Built DataFrame (after preprocessing): shape={df.shape}")

        # Generate dataset ID
        dataset_id = str(uuid.uuid4())

        # Get preview rows
        preview_rows = df.head(max_preview_rows).values.tolist()

        return (dataset_id, df, preview_rows)

    except Exception as e:
        logger.exception(f"Error building DataFrame: {file_path}, sheet={sheet_name}")
        raise

